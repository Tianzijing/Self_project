import datetime
from random import choice
from time import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

addr = r"C:\Users\Administrator\Desktop\test\ThinkSNS_registered.xlsx"
wb = Workbook()
# ws = wb.create_sheet()
ws = wb[wb.sheetnames[0]]
# ws.append(['TIME', 'TITLE', 'A-Z'])

for i in range(500):
    # TIME = datetime.datetime.now().strftime("%H:%M:%S")
    EMAIL = str(time())[:9] + str(i+100) + "@147.com"
    # A_Z = get_column_letter(choice(range(1, 50)))
    PASSWD = "1234546"
    PASSWD_r = "1234546"
    NAME = get_column_letter(choice(range(1000, 5000))) + str(choice(range(1000, 5000)))
    SIX = choice([1, 0])
    ws.append([NAME, PASSWD, PASSWD_r, EMAIL])

row_max = ws.max_row
con_max = ws.max_column
for j in ws.rows:
    for n in j:
        print(n.value, end="\t")
    print()
wb.save(addr)
