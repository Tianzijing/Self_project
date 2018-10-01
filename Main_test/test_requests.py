from unittest import TestCase

import requests
import datetime
from random import choice
from time import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# response = requests.get('http://www.baidu.com')
# print("状态码", response.status_code)	# 打印状态码
# print("请求URL", response.url)			# 打印请求url
# print("头信息", response.headers)		# 打印头信息
# print("cookie", response.cookies)		# 打印cookie信息
# print("文本形式  ", response.text)		# 以文本形式打印网页源码
# print("字节流形式", response.content)	# 以字节流形式打印

# response = requests.get('http://httpbin.org/get')
# # # response = requests.get('http://baiodu.com')
# print(response.text)
# # print(response.json())

# data = {'name': 'tom', 'age': '22'}
# response = requests.post('http://httpbin.org/post', data=data)
# print(response.text)

# response = requests.post('http://www.baidu.com')
# print(response.cookies)
# print(type(response.cookies))
# for k, v in response.cookies.items():
#     print(k+':'+v)

# session = requests.Session()
# session.get('http://httpbin.org/cookies/set/number/12345')
# response = session.get('http://httpbin.org/cookies')
# print(response.text)

# response = requests.get("https://www.zhihu.com")

# headers = {
#     # "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36"
#     # "User-Agent": "User-Agent:Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50"
#     "User-Agent": "User-Agent:Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)"
# }
# response = requests.get("https://www.zhihu.com", headers=headers)
# # print(response.apparent_encoding)
# response.encoding = response.apparent_encoding
# print(response.text)

addr = r"C:\Users\Administrator\Desktop\test\requests.xlsx"
wb = load_workbook(addr)
ws = wb[wb.sheetnames[0]]

row_max = ws.max_row
con_max = ws.max_column
for j in ws.rows:
    # print("开始使用浏览器：{} 发送请求 {}".format(j[0].value, j[1].value), end="\t")
    print("开始使用浏览器：{} 发送请求".format(j[0].value), end="\t")
    headers = {
        "User-Agent": j[1].value
    }
    response = requests.get("https://www.zhihu.com", headers=headers)
    # 获取自身编码格式，用于解码
    response.encoding = response.apparent_encoding
    http_test = response.text
    try:
        try:
            TestCase.assertIn(TestCase(), "知乎 - 有问题上知乎", http_test)
            print("成功登陆-断言成功！")
        except:
            TestCase.assertIn(TestCase(), "你正在使用的浏览器版本过低，将不能正常浏览和使用知乎。", http_test)
            print("提醒更新-断言成功！")
    except:
        print("断言失败！")
wb.save(addr)
